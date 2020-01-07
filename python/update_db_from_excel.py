# invenage uses an excel file as primary gui for adding items to database
# this script read the excel file and add items to the database
# it is very important to check for data integrity before adding to db
# ---------------------------- Setup ------------------------------------------
import sys, os, pandas as pd #import
import datetime
from openpyxl import load_workbook
inv_data_path = os.path.join(os.path.expanduser('~'),'invenage_data') 
sys.path.append(inv_data_path)
#build the config_dict
from python_conf import create_config_dict
config_dict = create_config_dict()
# add pythonPath
sys.path.append(os.path.join(config_dict['app_path'],'python'))
import python_func as inv
app_lang = config_dict['app_lang'] # get app_lang

# read the database
db_engine = inv.create_db_engine(config_dict)
conn = inv.db_open(config_dict,db_engine)
localisation = pd.read_sql_query(
        'select * from localisation where app_lang = "'+app_lang+'"',conn)
max_cust_id = pd.read_sql_query(
        'select max(customer_id) as value from customer_info',conn)
vendor_info = pd.read_sql_query('select * from vendor_info',conn)
warehouse_info = pd.read_sql_query('select * from warehouse_info',conn)
product_info = pd.read_sql_query('select * from product_info',conn)
packaging = pd.read_sql_query('select * from packaging',conn)
conn.close()

# error handling
error_file = config_dict['error_log']
msg_dict = inv.create_dict(config_dict,db_engine,'msg_dict')

# function to clear sheet
def clear_exceldb_sheet(update_file,sheet_name):
    wb = load_workbook(filename = update_file)
    ws = wb[sheet_name]
    for r in range(2,200):
        for c in range (1,30):
            ws.cell(row=r,column=c).value = ''
    wb.save(update_file)

# ------------------------------- Main ----------------------------------------
#update file location
update_file = config_dict['db_update_form']
#manupulate the localisation table to create a actual_col_name_to_label dict
acntl_dict = {}
for i in range(0,len(localisation)):
    acntl_dict[localisation.actual[i]] = localisation.label[i]

# -------------------- process add customer -----------------------------------
customer_sheet_name = config_dict['add_customer_sheetname']
append_cust_info = pd.read_excel(update_file,sheet_name = customer_sheet_name,
                                 dtype=str)
append_cust_info = append_cust_info.rename(columns=acntl_dict)

# will not add customer without a name
append_cust_info = append_cust_info[append_cust_info.customer_name.notnull()]
append_cust_info = append_cust_info[append_cust_info.customer_name!='']

# add customer id
append_cust_info['customer_id'] = max_cust_id.value[0]+1
append_cust_info.loc[
        append_cust_info.customer_tfn.isnull(),'customer_tfn'] = ''
# Remove all spaces in tfn
append_cust_info.customer_tfn = append_cust_info.customer_tfn.str.replace(
        ' ','')

# writing to database
conn = inv.db_open(config_dict, db_engine)
if len(append_cust_info)>0:
    append_cust_info.to_sql(
            'customer_info',conn,index=False,if_exists='append')
# conn.commit()
conn.close()
clear_exceldb_sheet(update_file,customer_sheet_name)

# ----------------------- process add product ---------------------------------
prod_sheet_name = config_dict['add_prod_sheetname']
tmp = pd.read_excel(update_file,sheet_name = prod_sheet_name)
tmp = tmp.rename(columns=acntl_dict)
# lower the unit
tmp.ordering_unit = tmp.ordering_unit.str.lower()
# remove whitespaces
tmp.ordering_unit = tmp.ordering_unit.str.strip()
tmp.prod_code = tmp.prod_code.str.replace(' ','')


tmp.ref_smn = tmp.ref_smn.astype(str)
# remove entries with invalid ordering unit, null prod_code, null warehouse
tmp = tmp[tmp.ordering_unit!='']
tmp = tmp[tmp.ordering_unit.notnull()]
tmp = tmp[tmp.prod_code.notnull()]
tmp = tmp[tmp.prod_code!='']
tmp = tmp[tmp.warehouse.notnull()]
tmp = tmp[tmp.warehouse!='']


# need to carefully check this sheet for integrity
tmp = pd.merge(tmp,vendor_info[['vendor','vendor_id']], how='left')
tmp = tmp[tmp.vendor_id.notnull()]
tmp = pd.merge(tmp,warehouse_info[['warehouse','warehouse_id']], how='left')
tmp = tmp[tmp.warehouse_id.notnull()]


# if there is an error. launch the log and shut down the script to protect
# data integrity
if (len(tmp[tmp.warehouse_id.isnull()])>0):
    inv.write_log(error_file,msg_dict['unknown_warehouse'])
    tmp.warehouse[tmp.warehouse_id.isnull()].to_csv(
            error_file,index=False,sep='\t',mode='a')
    inv.launch_file(error_file)
    raise RuntimeError('check error log')

if (len(tmp[tmp.vendor_id.isnull()])>0):
    inv.write_log(error_file,msg_dict['unknown_vendor'])
    tmp.vendor[tmp.vendor_id.isnull()].to_csv(
            error_file,index=False,sep='\t',mode='a')
    inv.launch_file(error_file)
    raise RuntimeError('check error log')
    
# otherwise prepare the final output
tmp['type'] = ''
tmp['import_license_exp'] = ''
tmp['updated_date'] = format(datetime.datetime.today(),'%d%m%y')
tmp['prod_group'] = ''
tmp['active'] = 1


append_pkg = tmp.copy()
append_pkg = append_pkg[['ordering_unit','prod_code','updated_date']]
append_pkg['unit']=append_pkg.ordering_unit
append_pkg.unit = append_pkg.unit.str.lower()
append_pkg['last_updated'] = append_pkg.updated_date
append_pkg['units_per_pack'] = 1 
# check for db existence and prepare output
append_pkg = inv.check_exists(
        append_pkg,packaging,['prod_code','unit'])
append_pkg = append_pkg[append_pkg.exist.isnull()]

append_pkg = append_pkg[['unit','units_per_pack','prod_code','last_updated']]

# check for db eistence and prepare output
tmp = inv.check_exists(tmp,product_info,['vendor','ref_smn'])
tmp = tmp[tmp.exist.isnull()]
tmp = tmp[['prod_code', 'name', 'vendor', 'ref_smn', 'type', 'packaging_str',
          'import_license_exp', 'updated_date', 'prod_group', 
          'warehouse_id', 'active']]
append_prod_info = tmp.copy()

conn = inv.db_open(config_dict, db_engine)
if len(append_prod_info)>0:
    append_prod_info.to_sql(
            'product_info',conn,index=False,if_exists='append')
    append_pkg.to_sql('packaging',conn,index=False,if_exists='append')
# conn.commit()
conn.close()
clear_exceldb_sheet(update_file,prod_sheet_name)

# ------------------------ process add packaging ------------------------------
pkg_sheet_name = config_dict['add_pkg_sheetname']
append_pkg_info = pd.read_excel(update_file,sheet_name = pkg_sheet_name)
append_pkg_info = append_pkg_info.rename(columns=acntl_dict)
append_pkg_info =append_pkg_info[append_pkg_info.prod_code.notnull()]
append_pkg_info['last_updated'] = format(datetime.datetime.now(),'%d%m%y')

# convert units_per_pack to number, removing invalid entries
append_pkg_info.units_per_pack = pd.to_numeric(
        append_pkg_info.units_per_pack,errors='coerce')
append_pkg_info = append_pkg_info[append_pkg_info.units_per_pack.notnull()]

# check if the prod_code is valid
append_pkg_info = pd.merge(
        append_pkg_info,product_info[['prod_code','name']],how='left')
append_pkg_info = append_pkg_info[append_pkg_info.prod_code.notnull()]

# clean up and prepare output
append_pkg_info.unit = append_pkg_info.unit.str.lower()
append_pkg_info = inv.check_exists(append_pkg_info,
                                   packaging,['prod_code','unit'])
append_pkg_info = append_pkg_info[append_pkg_info.exist.isnull()]
append_pkg_info = append_pkg_info[['unit','units_per_pack','prod_code',
                                   'last_updated']]
# writing to database
conn = inv.db_open(config_dict, db_engine)
if len(append_pkg_info)>0:
    append_pkg_info.to_sql(
        'packaging',conn,index=False,if_exists='append')
conn.close()
clear_exceldb_sheet(update_file,pkg_sheet_name)

if (config_dict['db_type'] == 'MariaDB'):
    db_engine.dispose()